from common.environment_variable import EnvironmentVariable
from common.style import Colour
from google import genai
from io import BytesIO
from openpyxl.styles import Alignment
from pathlib import Path
from peters_picturehouse_events.classes.movie_data_capture import MovieDataCapture
from peters_picturehouse_events.classes.movie_image import MovieImage
from PIL import Image
from requests.auth import HTTPBasicAuth
from common.messager import Messager

# import base64
import datetime
import json
import openpyxl
import os
import re
import requests
import urllib.parse
import utils.date_utils as du
import utils.device_utils as deu
import utils.file_utils as fu
import utils.string_utils as su
import peters_picturehouse_events.config as config


class MovieAutoProcessor:
    # def __init__(self, event_date_str: str, movie_name: str, release_year: str):
    def __init__(self, movie_data: MovieDataCapture):
        deu.clear_console()
        self.movie_data = movie_data

        self.event_date_str = movie_data.event_date
        self.search_name = movie_data.movie_title
        self.search_year = movie_data.movie_release_year
        self.omdb_api_key = EnvironmentVariable("OMDB_API_KEY", "string", False).value

        # Initialise empty member variables
        self.title = None
        self.release_year = None
        self.certification = None
        self.genre = None
        self.director = None
        self.starring = None
        self.plot = None
        self.poster_url = None
        self.imdbID = None
        self.summary = None
        self.running_time_raw = None
        self.running_time = None

        # Get templates
        self.listing_template = Path(config.TEMPLATE_FOLDER, "listing_template.html.txt")
        self.event_template = Path(config.TEMPLATE_FOLDER, "event_template.html.txt")

        # Automatically run the pipeline upon instantiation
        self._process_movie()

    def _get_formatted_filename(self) -> str:
        """
        # Translates the Excel formula logic into Python:
        """
        # LOWER and TRIM (strip leading/trailing whitespace, collapse multiple spaces to single)
        self.search_name = self.search_name.strip()
        clean_text = " ".join(self.search_name.lower().split())

        # SUBSTITUTE spaces with hyphens
        hyphenated = clean_text.replace(" ", "-")

        # REGEXREPLACE keeping only lowercase letters, numbers, and hyphens
        formatted_name = re.sub(r"[^a-z0-9\-]", "", hyphenated)

        # Clean up any potential double hyphens created by special characters
        formatted_name = re.sub(r"-+", "-", formatted_name)

        if self.search_year != "":
            return f"{formatted_name}_{self.search_year}.json"
        else:
            return f"{formatted_name}.json"

    def _process_movie(self):
        """Internal pipeline controller."""
        # 1. Check for OMDB data, reuse if found
        filename = self._get_formatted_filename()
        omdb_file_path = os.path.join(config.OMDB_FOLDER, filename)
        omdb_data = fu.load_json_if_exists(omdb_file_path)

        # 2. Convert date to required format(s)
        self.event_date = du.convert_string_to_date(self.event_date_str)
        self.date_formatted = du.convert_date_to_custom_string(self.event_date)

        # 3. Fetch OMDB data
        if not omdb_data:
            Messager(f"Getting fresh data from OMDB for movie {Colour.CYAN}{self.search_name}{Colour.RESET}", "level2")
            omdb_data = self._fetch_omdb_data()
            if not omdb_data:
                Messager(f"Movie '{Colour.CYAN}{self.search_name}{Colour.RESET}' not found on OMDB. Please check spelling and re-run.", "Error", True)
                return

        # 4. Store OMDB data into class member variables
        self.title = omdb_data.get("Title")
        self.release_year = omdb_data.get("Year")
        self.certification = omdb_data.get("Rated")
        self.genre = omdb_data.get("Genre")
        self.director = omdb_data.get("Director")
        self.starring = omdb_data.get("Actors")
        self.plot = omdb_data.get("Plot")
        self.poster_url = omdb_data.get("Poster")
        self.imdbID = omdb_data.get("imdbID")
        self.running_time_raw = omdb_data.get("Runtime")

        self.title_sanitised = su.sanitise_string(self.title)
        self.imdb_link = f"https://www.imdb.com/title/{self.imdbID}/"
        self.year_added, self.month_added = du.get_current_year_and_month()
        self.running_time = su.format_runtime(self.running_time_raw)

        # Verify essential data exists
        if not self.imdbID:
            print("Failed to acquire valid IMDb ID.")
            return

        # 5. Handle poster download and smart cropping
        if self.poster_url and self.poster_url != "N/A":
            self._download_and_process_poster()

        # 6. Handle Gemini summary check and generation
        self._get_or_generate_summary()

        # 7. Save final class output
        self._save_debug_json()

        # 8. Write back to Excel
        self._write_back_to_excel()

        # 9. (Future) Post to WordPress via REST API
        self.insert_event_via_api()

        # 10. Generate the listing
        self._generate_listing()

    def insert_event_via_api(self):
        """
        Get the template that contains the summary (body) of the event
        prior to insertion of data to replace placeholders
        """

        Messager(f"Creating event for movie {self.title} on WordPress", "l2")
        if self.event_exists(self.title, self.event_date):
            print(f"Event {Colour.CYAN}{self.title}{Colour.RESET} already exists on {self.event_date_str}. Skipping event creation.")
            return

        # Check if the image exists in Wordpress before attempting to create the event
        movie_image = MovieImage(self.title, self.title_sanitised)
        image_exists_in_wordpress, image_id, source_url = movie_image.check_image_exists_in_wordpress()

        if not image_exists_in_wordpress:
            movie_image.find_and_upload_local_image()
        else:
            # Process the source_url return, needed for the image link in the event and the listing.
            self.year_added, self.month_added = su.get_date_folders_from_url(source_url)

        # print(f"Creating event {Colour.CYAN}{self.title}{Colour.RESET} for {self.event_date_str}.")

        # Get the HTML template
        event_template: str = os.path.join(config.TEMPLATE_FOLDER, "event_template.html.txt")

        # Open the template
        with open(event_template, "r", encoding="utf-8") as file:
            template_content = file.read()
            # Make replacements
            self.description = self.replace_placeholders(template_content)

            event_data = {
                "title": f"Peter's Picturehouse: {self.title}",
                "description": self.description,
                "status": "publish",
                "start_date": self.event_date.strftime("%Y-%m-%d 19:00:00"),
                "end_date": self.event_date.strftime("%Y-%m-%d 21:30:00"),
                "cost": "5.00",
                "categories": [70],
                "venue": 840,
                "organizer": 4732,
                "show_map": True,
                "show_map_link": True,
            }

            response = requests.post(
                config.EVENTS_ENDPOINT,
                json=event_data,
                auth=HTTPBasicAuth(
                    config.WORDPRESS_USERNAME,
                    config.WORDPRESS_APPLICATION_KEY,
                ),
            )

            if response.status_code == 201:
                # result = response.json()
                # print(f"{Colour.GREEN}Successfully created event! ID: {result.get('id')}{Colour.RESET}\n")
                Messager(f"{Colour.GREEN}Event {self.title} successfully created.{Colour.RESET}", "normal")
            else:
                Messager(f"Failed to create event. Status code: {response.status_code}", "error")
                # print(f"Failed to create event. Status code: {response.status_code}")

    def event_exists(self, title, start_date_obj):
        """
        Checks if an event with the same title exists on the same start date.
        start_date_obj should be a python date or datetime object.
        """
        title_sanitised = su.sanitise_string(self.title)
        # Create a window for the entire day
        day_start = start_date_obj.strftime("%Y-%m-%d 00:00:00")
        day_end = start_date_obj.strftime("%Y-%m-%d 23:59:59")

        params = {"start_date": day_start, "end_date": day_end, "per_page": 50}

        response = requests.get(
            config.EVENTS_ENDPOINT,
            params=params,
            auth=HTTPBasicAuth(
                config.WORDPRESS_USERNAME,
                config.WORDPRESS_APPLICATION_KEY,
            ),
        )

        if response.status_code == 200:
            existing_events = response.json().get("events", [])

            # Now we look for the specific title match within that day
            for event in existing_events:
                wp_title = event.get("title", "")
                wp_title = su.sanitise_string(wp_title)
                if title_sanitised in wp_title:
                    return True

        return False

    def _generate_listing(self):
        """
        Generates a listing HTML file that can then be copied and pasted into WordPress
        Other, more technically appropriate approaches have been explore but are not currently
        possible in the Netwise environment, which has an old version of ACF which cannot
        accommodate the necessary custom fields.
        """
        Messager(f"Generating listing for movie {self.title}", "l2")
        listing_html = fu.read_file_content(self.listing_template)
        listing_html = self.replace_placeholders(listing_html)
        listing_file = Path(config.LISTING_FOLDER, f"{self.title_sanitised}.html.txt")
        with open(listing_file, "w", encoding="utf-8") as f:
            f.write(listing_html)
        Messager("Listing HTML created. This needs to be copied into the Peter's Picturehouse page on WordPress.", "dim")

    def _write_back_to_excel(self):
        """
        This function takes a newly entered movive, checks if it exists already in the Excel master
        and if not, it repopulates it with the data that has been extracted from the OMDB API and
        from Gemini.
        """

        Messager(f"Writing data for movie {Colour.CYAN}{self.title}{Colour.RESET} back to the master Excel document.", "level2")

        excel_filename = "picture-house-data.xlsx"
        excel_filepath = os.path.join(config.INPUT_FOLDER, excel_filename)
        sheet_name = "movies"

        # 1. Open the xlsx file
        workbook = openpyxl.load_workbook(excel_filepath)
        sheet = workbook[sheet_name]

        # 2. Check if a row with the same date already exists
        target_date = self.event_date
        target_row = None

        for row in range(1, sheet.max_row + 1):
            cell_value = sheet[f"A{row}"].value

            # Normalise openpyxl values to datetime.date objects for comparison
            if isinstance(cell_value, datetime.datetime):
                cell_date = cell_value.date()
            elif isinstance(cell_value, datetime.date):
                cell_date = cell_value
            else:
                cell_date = None

            if cell_date == target_date:
                target_row = row
                break

        # Determine if we are overwriting an existing row or appending a new one
        if target_row is not None:
            display_date = target_date.strftime("%d/%m/%Y")
            user_choice = (
                input(
                    f"\nA row with the date {Colour.CYAN}{display_date}{Colour.RESET} already exists in the Excel at row {target_row}. {Colour.CYAN}Overwrite? (yes / no){Colour.RESET}?"
                )
                .strip()
                .lower()
            )
            if user_choice not in ["yes", "y"]:
                # print("Operation cancelled. No changes made.")
                workbook.close()
                return

            Messager("\nRow updated in Excel", "dim")
            active_row = target_row
        else:
            active_row = sheet.max_row + 1
            Messager("\nNew row written to Excel", "dim")

        # 3. Enter the data into the determined row
        # Date cell
        cell = sheet[f"A{active_row}"]
        cell.value = target_date
        cell.number_format = "DD/MM/YYYY"

        # Other cells
        sheet[f"B{active_row}"] = self.title
        sheet[f"C{active_row}"] = self.imdb_link
        sheet[f"D{active_row}"] = self.summary
        sheet[f"E{active_row}"] = self.starring
        sheet[f"F{active_row}"] = self.certification
        sheet[f"G{active_row}"] = self.running_time
        sheet[f"H{active_row}"] = self.release_year
        sheet[f"I{active_row}"] = self.title_sanitised
        sheet[f"K{active_row}"] = self.year_added
        sheet[f"L{active_row}"] = self.month_added

        alignments = {
            "A": Alignment(horizontal="left", vertical="top"),
            "B": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "C": Alignment(horizontal="left", vertical="top"),
            "D": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "E": Alignment(horizontal="left", vertical="top", wrap_text=True),
            "F": Alignment(horizontal="center", vertical="top"),
            "G": Alignment(horizontal="center", vertical="top"),
            "H": Alignment(horizontal="center", vertical="top"),
            "I": Alignment(horizontal="left", vertical="top"),
            "J": Alignment(horizontal="center", vertical="top"),
            "K": Alignment(horizontal="center", vertical="top"),
            "L": Alignment(horizontal="center", vertical="top"),
        }
        for cell_ref, alignment in alignments.items():
            cell = sheet[f"{cell_ref}{active_row}"]
            cell.alignment = alignment

        # 4. Save the xlsx file with no collateral damage
        workbook.save(excel_filepath)
        workbook.close()

    def _fetch_omdb_data(self) -> dict:
        """Formats querystring variables, retrieves data from OMDB, and saves a local copy."""
        # URL encode the movie name safely to handle spaces and special characters
        encoded_name = urllib.parse.quote_plus(self.search_name)
        url = f"http://www.omdbapi.com/?t={encoded_name}&y={self.search_year}&apikey={self.omdb_api_key}"

        response = requests.get(url)
        response.raise_for_status()
        data = response.json()

        if data.get("Response") == "True":
            # Ensure the resources/omdb directory exists
            os.makedirs(config.OMDB_FOLDER, exist_ok=True)

            # Generate the specific filename based on your formatting rules
            filename = self._get_formatted_filename()
            omdb_file_path = os.path.join(config.OMDB_FOLDER, filename)

            # Save the raw OMDB JSON output locally
            with open(omdb_file_path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
            Messager(f"OMDB response cached locally as: {filename}", "dim")

            return data
        return {}

    def _download_and_process_poster(self):
        Messager("Processing movie poster", "l1")
        """Downloads poster image and crops/resizes it to 320x474 without stretching."""
        OVERWRITE_IMAGES = EnvironmentVariable("OVERWRITE_IMAGES", "bool", False).value
        os.makedirs(config.IMAGES_FOLDER, exist_ok=True)
        poster_filename = f"{self.title_sanitised}.webp"
        poster_path = os.path.join(config.IMAGES_FOLDER, poster_filename)
        if os.path.exists(poster_path):
            if not OVERWRITE_IMAGES:
                Messager("Poster image already exists - using original", "dim")
                return
        try:
            # print("Getting image")
            response = requests.get(self.poster_url)
            response.raise_for_status()

            img = Image.open(BytesIO(response.content))

            target_w, target_h = 320, 474
            orig_w, orig_h = img.size

            scale_w = target_w / orig_w
            scale_h = target_h / orig_h
            scale = max(scale_w, scale_h)

            new_w = int(orig_w * scale)
            new_h = int(orig_h * scale)
            img_resized = img.resize((new_w, new_h), Image.Resampling.LANCZOS)

            left = (new_w - target_w) / 2
            top = (new_h - target_h) / 2
            right = (new_w + target_w) / 2
            bottom = (new_h + target_h) / 2

            img_cropped = img_resized.crop((left, top, right, bottom))

            os.makedirs(config.IMAGES_FOLDER, exist_ok=True)
            poster_filename = f"{self.title_sanitised}.webp"
            poster_path = os.path.join(config.IMAGES_FOLDER, poster_filename)
            img_cropped.save(poster_path, "WEBP", quality=45)
            Messager(f"Poster image saved as {Colour.CYAN}{poster_filename}{Colour.RESET}.\n", "dim")

        except Exception as e:
            Messager(f"Failed to process poster image: {e}", "error", True)

    def _get_or_generate_summary(self) -> str:
        """
        Retrieves cached summary JSON locally, or hits Gemini API to generate it.
        """
        Messager("Getting movie summary", "l1")
        self.summary = ""

        # Look for a locally downloaded file with the summary in it
        summary_filename = f"{self.title_sanitised}.json"
        summary_filepath = os.path.join(config.SUMMARY_FOLDER, summary_filename)
        if os.path.exists(summary_filepath):
            Messager(f"Using cached summary at: {summary_filename}", "dim")
            with open(summary_filepath, "r", encoding="utf-8") as f:
                cached_data = json.load(f)
                self.summary = cached_data.get("summary", "")
                return "Success"

        os.makedirs(config.SUMMARY_FOLDER, exist_ok=True)

        if not os.environ.get("GEMINI_API_KEY"):
            Messager("GEMINI_API_KEY environment variable not set.", "error", True)
            return
        client = genai.Client()

        # This is the prompt that is sent to Gemini to get the summary of the movie
        prompt = (
            f"Provide a concise summary of the movie '{self.title}' ({self.release_year}). "
            f"The summary MUST be between 110 and 140 words in length. "
            f"Do not include any spoilers. "
            f"Use UK spelling. "
            f"Return the response in a clean JSON format matching this schema: "
            f'{{"summary": "your text summary here"}}'
        )

        response = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=prompt,
            config={"response_mime_type": "application/json"},
        )

        try:
            gemini_json = json.loads(response.text)
            self.summary = gemini_json.get("summary", "").strip()

            with open(summary_filepath, "w", encoding="utf-8") as f:
                json.dump(gemini_json, f, indent=2, ensure_ascii=False)

            Messager(f"Movie summary saved as {summary_filename}", "dim")

            return "Success"
        except Exception as e:
            Messager(f"Error handling Gemini JSON structure: {e}.", "error", True)
            self.summary = ""
            return response.text

    def _save_debug_json(self):
        """
        Compiles class instance state into a JSON dictionary and exports it.
        """
        os.makedirs(config.DEBUG_FOLDER, exist_ok=True)
        output_filename = f"{self.title_sanitised}.json"
        output_file = os.path.join(config.DEBUG_FOLDER, output_filename)

        data_to_export = {
            "title": self.title,
            "year": self.release_year,
            "rating": self.certification,
            "genre": self.genre,
            "director": self.director,
            "actors": self.starring,
            "plot": self.plot,
            "poster_url": self.poster_url,
            "imdbID": self.imdbID,
            "summary": self.summary,
        }

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data_to_export, f, indent=2, ensure_ascii=False)
            # print(f"Full movie class data exported safely to: {output_filename}")

    def replace_placeholders(self, s):
        try:
            s = s.replace("{title}", self.title)
            s = s.replace("{year_added}", self.year_added)
            s = s.replace("{month_added}", self.month_added)
            s = s.replace("{title_formatted}", self.title_sanitised)
            s = s.replace("{imdb_link}", self.imdb_link)
            s = s.replace("{date_formatted}", self.date_formatted)
            s = s.replace("{certification}", self.certification)
            s = s.replace("{release_year}", self.release_year)
            s = s.replace("{summary}", self.summary)
            s = s.replace("{starring}", self.starring)
            s = s.replace("{running_time}", self.running_time)
        except Exception as e:
            print(e)
            pass
        return s
